# One class to for all your x2latex conversion needs


# Needed imports ---------------------------------------------

# imports from this package
from .texformat import texfmt # converter of data types to latex
from .xl_utils import wb2list, maketable, dftable, col2row # reads data table
# helper function to convert images and make latex figures
from .img_utils import get_files, graphicx_compatible, mk_img_fig
from .csv_utils import csv2table # convert a csv file to a 2D list



# Class --------------------------------------------------------

'''
a class that converts various data types to latex representations

Can convert:
> Excel via openpyxl or pandas (pyxl is prefered)
> Images
> Text files
> raw data types

for all text, a single \ means that str parser will add
$ $ around the mathematical phrase
\\ escapes that and leaves only a single \ for non-math
latex commands
'''
class totex:

    def __init__(self, **kwargs):

        # creates a texfmt instance
        # this controls all formatting
        self.fmt = texfmt(**kwargs)

        return


    # xl_to_tex takes a path to a spreadsheet and formats all the pages into a latex table
    # wraps the df_to_tex function
    def xl(self, file, sheets=None,**kwargs):
        '''
        Parameters:

        > FILE: is the path to the excel spreadsheet
        > SHEETS: a list of names of the workbook sheets to convert
        > BACKEND is either 'pyxl' or 'pandas'
        > KWARGS: are the keyword arguments passed to the latex
            table creation function
        '''
        # import for reading xl tables
        from openpyxl import load_workbook
        
        # importing file
        workbook = load_workbook(file, data_only=True) # openpyxl
            

        # Using default sheet settings (all sheets in workbook)
        if not sheets:
            sheets = workbook.sheetnames

        # converting a single sheet to latex
        if isinstance(sheets, str):
            # forcing sheets to be a list
            sheets = [sheets] 

        # converting all sheets to latex
        # analysing data on every page of the sheet seperately
        commands = []

        # iterating over all sheets
        for i in range(len(sheets)):

            # loading speadsheet data from a particular page
            table = wb2list(workbook[sheets[i]]) # xl_utils function

            # turns dataframe into latex table
            commands.append(maketable(table,formatter=self.fmt,**kwargs)) # xl_utils function

            # return a list of tables
            return commands


    # ---------------------------------------------------------------------------------

    # read from an excel spreadsheet with pandas
    # and convert it to a spreadsheet
    def xldf(self, file, sheets=[0]):
        '''
        Parameters:

        > FILE: a path to an xlsx file
        > SHEETS a list of sheet names to read, default is to read the first
            sheet.
        '''
        
        # reading backend
        from pandas import read_excel

        # stores all formated tex commands to return
        commands = []
        
        # importing file
        for i in sheets:

            # reading in this excel sheet
            wb = read_excel(file, sheet_name=i)

            # formating sheet and appending it
            commands.append(self.df(wb))

        return commands
        

    # ----------------------------------------------------------------------------------

    # puts together all the other functions and converts an image to tex commands
    def img(self, file_dir=None, images=None, convert=True,**kwargs):
        '''
        Parameters:

        > FILE_DIR: path to a directory that has images to be converted to
            latex commands
        > IMAGES: path to specific images to convert
        > CONVERT: decision to either convert images to .eps format or not
            This is the prefered latex format
        '''


        # if no images are given than all images in file_dir are used
        if not images and file_dir:
            # gets all images in KNOWN_EXT
            images = get_files(file_dir=file_dir)
        
        # creating the list of latex command to render the images
        img_command = []

        # accounting for if IMAGES argument is a single string
        if isinstance(images, str):
            images = [images]

        # iterates over every image,
        # checks compatibility and
        # creates a figure of that image
        for i in range(len(images)):

            # forceing images to be latex compatible
            images[i] = graphicx_compatible(images[i], make=convert)

            # if converting to eps failed, skip this step
            if not images[i]:
                img_command.append(None)
                continue
                

            # storing figures with images
            img_command.append(mk_img_fig(images[i], **kwargs))


        # returning latex commands. End of function
        return img_command


    # --------------------------------------------------------------------------------------

    # reads the contents of a text file and converts them to
    # latex commands. Very simple conversion (only math and simple
    # latex commands)
    def txt(self, file, savefile=None, **kwargs):
        '''
        Parameters

        > FILE: the path to the file to read as a sting
        > SAVEFILE: optional to create and save the formated text
            to a new file
        > KWARGS: passed to the texFormater class on creation
        '''

        # opening the given file in read mode
        # assumes a simple text file
        f = open(file)

        # getting text from file
        text = f.read()

        # closing the opened file after
        # it has been read
        f.close()

        # if a save file is given, then
        # the formated text is saved into this file
        if savefile and isinstance(savefile,str):

            # creating and opening file in write mode
            f1 = open(savefile,'w')

            # writing the formatted string to the save file
            f1.write(self.fmt.to_str(text))

            # closing new file
            f1.close()

            return
            

        # returning the formatte
        return self.fmt.to_str(text)


    # ------------------------------------------------------------------

    # converting csv files to latex
    def csv(self, file, **kwargs):
        '''
        Parameters:

        > FILE: the path to the csv file
        '''

        # getting csv data into table
        table = csv2table(file, **kwargs)

        # switch sublist representation
        table = col2row(table)

        # doing string formatting
        return maketable(table, formatter=self.fmt)
        
    # -------------------------------------------------------------------

    # method to covert a DataFrame to latex table
    def df(self, dataframe, dropna='all',**kwargs):
        '''
        Parameters:

        > DATAFRAME: a single data frame to convert to a latex table
            This does not use the builtin df.to_latex method
            but functions like the xl method in this class
        '''

        # returning table
        return dftable(dataframe, formatter=self.fmt,**kwargs)

    # -------------------------------------------------------------------

    # a wrapper over the texfmt to_str method
    def raw(self, thing):
        '''
        Parameters:

        > THING: a raw datatype or nan
        '''
        return self.fmt.to_str(thing)
        


























    
